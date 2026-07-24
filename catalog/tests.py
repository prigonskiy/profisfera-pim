"""
Тесты критических путей PIM.

Покрывают то, что молча ломается при правках и бьёт по данным/доступу:
  • присвоение внутреннего артикула sku (формат, уникальность, неизменность);
  • права API (аноним читает, но не пишет; сотрудник пишет);
  • обращение к товару и по slug, и по sku;
  • импорт xlsx: создание/обновление, сопоставление по ключам и ТИПОВАЯ
    предпроверка (мусор в булевой/числовой ячейке отклоняет строку);
  • статусы срока действия документов;
  • совместимость форматов экспорта и импорта (round-trip).

Запуск:  python manage.py test catalog
"""
from datetime import timedelta
from io import BytesIO
import os
import tempfile
from unittest import mock

from django.contrib.auth.models import User
from django.core.cache import cache
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase, override_settings
from django.urls import reverse
from django.utils import timezone
from openpyxl import Workbook
from rest_framework.test import APIClient

from catalog import catalog_io
from catalog.models import (
    Category,
    CategoryFilter,
    Characteristic,
    CharacteristicOption,
    Document,
    Product,
    ProductAttributeValue,
    SkuCounter,
)


def make_import_xlsx(id_row, data_rows):
    """Собрать xlsx нашего формата: строка 1 — идентификаторы, строка 2 — подписи, данные с 3-й."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Товары"
    ws.append(id_row)
    ws.append([f"({c})" for c in id_row])  # строка-подписи, импорт её игнорирует
    for row in data_rows:
        ws.append(row)
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


class SkuTests(TestCase):
    def setUp(self):
        self.cat = Category.objects.create(name="Материалы")

    def test_sku_assigned_format_and_sequence(self):
        p1 = Product.objects.create(name="Товар 1", category=self.cat)
        p2 = Product.objects.create(name="Товар 2", category=self.cat)
        # первый артикул = старт счётчика + 1, длиной 10 знаков
        self.assertEqual(p1.sku, str(SkuCounter.SKU_START + 1))
        self.assertEqual(len(p1.sku), 10)
        # счётчик монотонно растёт и не повторяется
        self.assertEqual(int(p2.sku), int(p1.sku) + 1)

    def test_sku_is_stable_on_resave(self):
        p = Product.objects.create(name="Товар", category=self.cat)
        original = p.sku
        p.name = "Товар (переименован)"
        p.save()
        p.refresh_from_db()
        self.assertEqual(p.sku, original)

    def test_sku_not_reused_after_delete(self):
        p1 = Product.objects.create(name="A", category=self.cat)
        first = p1.sku
        p1.delete()
        p2 = Product.objects.create(name="B", category=self.cat)
        self.assertNotEqual(p2.sku, first)
        self.assertGreater(int(p2.sku), int(first))


class ApiPermissionTests(TestCase):
    def setUp(self):
        cache.clear()  # сбросить историю троттлинга, чтобы тесты не влияли друг на друга
        self.cat = Category.objects.create(name="Материалы")
        self.product = Product.objects.create(name="Адгезив", category=self.cat)
        self.staff = User.objects.create_user("staff", password="x", is_staff=True)

    def test_anonymous_can_read_list_and_detail(self):
        client = APIClient()
        self.assertEqual(client.get("/api/products/").status_code, 200)
        self.assertEqual(client.get(f"/api/products/{self.product.slug}/").status_code, 200)

    def test_product_addressable_by_slug_and_sku(self):
        client = APIClient()
        by_slug = client.get(f"/api/products/{self.product.slug}/")
        by_sku = client.get(f"/api/products/{self.product.sku}/")
        self.assertEqual(by_slug.status_code, 200)
        self.assertEqual(by_sku.status_code, 200)
        self.assertEqual(by_slug.data["id"], by_sku.data["id"])

    def test_anonymous_cannot_write(self):
        client = APIClient()
        self.assertIn(client.post("/api/products/", {"name": "X"}).status_code, (401, 403))
        self.assertIn(client.patch(f"/api/products/{self.product.slug}/", {}).status_code, (401, 403))
        self.assertIn(client.delete(f"/api/products/{self.product.slug}/").status_code, (401, 403))

    def test_staff_can_write(self):
        client = APIClient()
        client.force_authenticate(user=self.staff)
        resp = client.patch(f"/api/products/{self.product.slug}/", {})
        self.assertEqual(resp.status_code, 200)

    def test_authenticated_non_staff_cannot_write(self):
        user = User.objects.create_user("plain", password="x", is_staff=False)
        client = APIClient()
        client.force_authenticate(user=user)
        self.assertIn(client.patch(f"/api/products/{self.product.slug}/", {}).status_code, (401, 403))


class DocumentStatusTests(TestCase):
    def _doc(self, **kwargs):
        return Document.objects.create(name="Док", **kwargs)

    def test_perpetual(self):
        self.assertEqual(self._doc(is_perpetual=True).status, Document.Status.PERPETUAL)

    def test_unknown_when_no_date(self):
        self.assertEqual(self._doc().status, Document.Status.UNKNOWN)

    def test_expired(self):
        d = self._doc(valid_until=timezone.localdate() - timedelta(days=1))
        self.assertEqual(d.status, Document.Status.EXPIRED)
        self.assertLess(d.days_left, 0)

    def test_expiring_soon(self):
        d = self._doc(valid_until=timezone.localdate() + timedelta(days=10))
        self.assertEqual(d.status, Document.Status.EXPIRING)
        self.assertEqual(d.days_left, 10)

    def test_valid(self):
        d = self._doc(valid_until=timezone.localdate() + timedelta(days=400))
        self.assertEqual(d.status, Document.Status.VALID)


class XlsxImportTests(TestCase):
    def setUp(self):
        self.cat = Category.objects.create(name="Материалы")
        self.bool_char = Characteristic.objects.create(
            name="Тестовый флаг", code="test_flag",
            type=Characteristic.Type.BOOLEAN, is_global=True,
        )

    def test_dry_run_creates_nothing_but_reports(self):
        f = make_import_xlsx(
            ["sku", "name", "category_slug"],
            [["", "Новый товар для dry-run", self.cat.slug]],
        )
        report = catalog_io.import_workbook(f, dry_run=True)
        self.assertEqual(report["created"], 1)
        self.assertFalse(report["applied"])
        self.assertEqual(Product.objects.filter(name="Новый товар для dry-run").count(), 0)  # откатилось

    def test_real_import_creates_product_with_sku(self):
        f = make_import_xlsx(
            ["sku", "name", "category_slug", "char:test_flag"],
            [["", "Адгезив Bond", self.cat.slug, "Да"]],
        )
        report = catalog_io.import_workbook(f, dry_run=False)
        self.assertEqual(report["created"], 1)
        self.assertTrue(report["applied"])
        p = Product.objects.get(name="Адгезив Bond")
        self.assertTrue(p.sku)  # внутренний артикул присвоен
        av = ProductAttributeValue.objects.get(product=p, characteristic=self.bool_char)
        self.assertIs(av.value_boolean, True)

    def test_update_by_sku(self):
        p = Product.objects.create(name="Старое имя", category=self.cat)
        f = make_import_xlsx(
            ["sku", "name", "category_slug"],
            [[p.sku, "Новое имя", self.cat.slug]],
        )
        report = catalog_io.import_workbook(f, dry_run=False)
        self.assertEqual(report["updated"], 1)
        p.refresh_from_db()
        self.assertEqual(p.name, "Новое имя")

    def test_garbage_in_boolean_is_rejected(self):
        """Кейс из практики: «цц» в булевой — строка должна отклоняться, а не молча проглатываться."""
        f = make_import_xlsx(
            ["sku", "name", "category_slug", "char:test_flag"],
            [["", "Битый товар", self.cat.slug, "цц"]],
        )
        report = catalog_io.import_workbook(f, dry_run=False)
        self.assertEqual(report["created"], 0)
        self.assertGreaterEqual(len(report["errors"]), 1)
        self.assertEqual(Product.objects.filter(name="Битый товар").count(), 0)

    def test_unknown_sku_is_error(self):
        f = make_import_xlsx(
            ["sku", "name", "category_slug"],
            [["9999999999", "Призрак", self.cat.slug]],
        )
        report = catalog_io.import_workbook(f, dry_run=False)
        self.assertEqual(report["created"], 0)
        self.assertEqual(report["updated"], 0)
        self.assertGreaterEqual(len(report["errors"]), 1)


class ExportRoundTripTests(TestCase):
    def setUp(self):
        self.cat = Category.objects.create(name="Материалы")
        self.product = Product.objects.create(
            name="Адгезив Bond Force", category=self.cat, manufacturer_sku="14926",
        )

    def test_export_structure(self):
        wb = catalog_io.build_export_workbook(Product.objects.all(), include_category_chars=False)
        ws = wb["Товары"]
        self.assertEqual(ws.cell(row=1, column=1).value, "sku")            # строка 1 — идентификаторы
        self.assertEqual(ws.cell(row=3, column=1).value, self.product.sku)  # данные с 3-й строки

    def test_export_then_import_is_consistent(self):
        """Файл из экспорта должен без ошибок читаться импортом — один и тот же формат."""
        wb = catalog_io.build_export_workbook(Product.objects.all(), include_category_chars=False)
        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        report = catalog_io.import_workbook(bio, dry_run=True)
        self.assertEqual(report["errors"], [])
        self.assertEqual(report["updated"], 1)
        self.assertEqual(report["created"], 0)


class ThrottleConfigTests(TestCase):
    """Троттлинг должен оставаться включённым в настройках API — защита от случайного удаления.

    Живое срабатывание 429 здесь не воспроизводим (поведение троттла под тест-клиентом
    DRF капризно и зависит от кэша/инстансов); проверяем сам факт настройки, а 429
    подтверждается вручную, напр.:  for i in $(seq 1 400); do curl -s -o /dev/null -w "%{http_code}\\n" <API>/api/products/; done | sort | uniq -c
    """

    def test_throttle_classes_and_rates_configured(self):
        from rest_framework.settings import api_settings
        from rest_framework.throttling import AnonRateThrottle, UserRateThrottle

        classes = list(api_settings.DEFAULT_THROTTLE_CLASSES)
        self.assertIn(AnonRateThrottle, classes)
        self.assertIn(UserRateThrottle, classes)
        rates = api_settings.DEFAULT_THROTTLE_RATES
        self.assertTrue(rates.get("anon"))
        self.assertTrue(rates.get("user"))


class CustomerVisibilityTests(TestCase):
    """Что видит покупатель: служебные характеристики скрыты, имя документа необязательно."""

    def setUp(self):
        cache.clear()
        self.cat = Category.objects.create(name="Материалы")
        self.product = Product.objects.create(name="Товар", category=self.cat)

    def test_hidden_characteristic_not_in_api(self):
        visible = Characteristic.objects.create(
            name="Цвет", code="color_vis", type=Characteristic.Type.TEXT,
            is_global=True, show_to_customer=True,
        )
        hidden = Characteristic.objects.create(
            name="Код 1С", code="code_1c", type=Characteristic.Type.TEXT,
            is_global=True, show_to_customer=False,
        )
        ProductAttributeValue.objects.create(
            product=self.product, characteristic=visible, value_text="белый")
        ProductAttributeValue.objects.create(
            product=self.product, characteristic=hidden, value_text="00-0001")

        resp = APIClient().get(f"/api/products/{self.product.slug}/")
        codes = [c["code"] for c in resp.data["characteristics"]]
        self.assertIn("color_vis", codes)       # обычная — видна
        self.assertNotIn("code_1c", codes)      # служебная — скрыта

    def test_document_name_is_optional(self):
        d = Document.objects.create(number="РЗН 2016/4080")  # без name
        self.assertEqual(d.name, "")
        self.assertTrue(Document.objects.filter(pk=d.pk).exists())


class StorefrontTriggerTests(TestCase):
    """Триггер пересборки витрины: безопасные пути без обращения к сети."""

    def test_not_configured_returns_error_without_network(self):
        from catalog.storefront import trigger_rebuild
        with mock.patch.dict(os.environ, {"GITHUB_DISPATCH_TOKEN": "", "STOREFRONT_REPO": ""}):
            ok, msg = trigger_rebuild()
        self.assertFalse(ok)
        self.assertIn("не настроена", msg.lower())

    def test_success_on_204(self):
        from catalog import storefront

        class _Resp:
            def __enter__(self):
                return self
            def __exit__(self, *a):
                return False
            def getcode(self):
                return 204

        env = {
            "GITHUB_DISPATCH_TOKEN": "dummy",
            "STOREFRONT_REPO": "owner/repo",
            "STOREFRONT_DISPATCH_EVENT": "rebuild",
        }
        with mock.patch.dict(os.environ, env), \
             mock.patch.object(storefront.urllib.request, "urlopen", return_value=_Resp()):
            ok, msg = storefront.trigger_rebuild()
        self.assertTrue(ok)


class AdminSkuUrlTests(TestCase):
    """Адресация карточки товара в админке: URL по sku, старые pk-ссылки тоже живут."""

    def setUp(self):
        self.cat = Category.objects.create(name="Материалы")
        self.product = Product.objects.create(name="Адгезив", category=self.cat)
        User.objects.create_superuser("admin_url", "a@example.com", "pw12345")
        self.client.force_login(User.objects.get(username="admin_url"))

    def test_change_url_uses_sku(self):
        url = reverse("admin:catalog_product_change", args=[self.product.sku])
        self.assertIn(self.product.sku, url)
        resp = self.client.get(url)
        self.assertEqual(resp.status_code, 200)

    def test_old_pk_url_still_works(self):
        url = f"/admin/catalog/product/{self.product.pk}/change/"
        resp = self.client.get(url)
        self.assertEqual(resp.status_code, 200)

    def test_changelist_links_point_to_sku(self):
        resp = self.client.get(reverse("admin:catalog_product_changelist"))
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, f"/product/{self.product.sku}/change/")


@override_settings(MEDIA_ROOT=tempfile.mkdtemp())
class DocumentProductsWidgetTests(TestCase):
    """Привязка документа к товарам одним списком (DocumentAdminForm)."""

    def setUp(self):
        self.cat = Category.objects.create(name="Материалы")
        self.p1 = Product.objects.create(name="Композитная краска A2", category=self.cat)
        self.p2 = Product.objects.create(name="Композитная краска A3", category=self.cat)

    def test_initial_loads_existing_products(self):
        from catalog.admin import DocumentAdminForm
        doc = Document.objects.create(name="РУ")
        doc.products.set([self.p1])
        form = DocumentAdminForm(instance=doc)
        self.assertIn(self.p1, list(form.fields["products"].initial))
        self.assertNotIn(self.p2, list(form.fields["products"].initial))

    def test_save_persists_selected_products(self):
        from catalog.admin import DocumentAdminForm
        pdf = SimpleUploadedFile("ru.pdf", b"%PDF-1.4 test", content_type="application/pdf")
        form = DocumentAdminForm(
            data={
                "name": "РУ на серию",
                "doc_type": Document.DocType.REGISTRATION,
                "number": "",
                "issuing_authority": "",
                "products": [self.p1.pk, self.p2.pk],
            },
            files={"file": pdf},
        )
        self.assertTrue(form.is_valid(), form.errors)
        doc = form.save()
        self.assertEqual(
            set(doc.products.values_list("pk", flat=True)),
            {self.p1.pk, self.p2.pk},
        )


class CategoryFilterTests(TestCase):
    """Категорийные фильтры: наследование, валидация типа→вида, эндпоинт конфига."""

    def setUp(self):
        cache.clear()
        self.parent = Category.objects.create(name="Брекеты")
        self.child = Category.objects.create(name="Металлические", parent=self.parent)
        # числовая характеристика на родителе
        self.paz = Characteristic.objects.create(name="Паз", code="paz", type=Characteristic.Type.NUMBER, unit="дюйм")
        self.paz.categories.add(self.parent)
        # select-характеристика на ребёнке
        self.system = Characteristic.objects.create(name="Система", code="system", type=Characteristic.Type.SINGLE_SELECT)
        self.system.categories.add(self.child)
        CharacteristicOption.objects.create(characteristic=self.system, value="Roth")
        CharacteristicOption.objects.create(characteristic=self.system, value="MBT")

    def test_child_inherits_parent_filter(self):
        CategoryFilter.objects.create(category=self.parent, characteristic=self.paz,
                                      display=CategoryFilter.Display.NUMBER_RANGE)
        CategoryFilter.objects.create(category=self.child, characteristic=self.system,
                                      display=CategoryFilter.Display.SELECT_CHECKBOX)
        codes = [f.characteristic.code for f in self.child.effective_filters()]
        self.assertIn("paz", codes)      # унаследован от родителя
        self.assertIn("system", codes)   # собственный

    def test_child_overrides_parent_for_same_characteristic(self):
        CategoryFilter.objects.create(category=self.parent, characteristic=self.paz,
                                      display=CategoryFilter.Display.NUMBER_RANGE)
        CategoryFilter.objects.create(category=self.child, characteristic=self.paz,
                                      display=CategoryFilter.Display.NUMBER_BUCKETS)
        eff = {f.characteristic.code: f.display for f in self.child.effective_filters()}
        self.assertEqual(eff["paz"], CategoryFilter.Display.NUMBER_BUCKETS)

    def test_clean_rejects_text_characteristic(self):
        from django.core.exceptions import ValidationError
        txt = Characteristic.objects.create(name="Заметка", code="note", type=Characteristic.Type.TEXT)
        cf = CategoryFilter(category=self.parent, characteristic=txt,
                            display=CategoryFilter.Display.SELECT_CHECKBOX)
        with self.assertRaises(ValidationError):
            cf.clean()

    def test_clean_rejects_display_type_mismatch(self):
        from django.core.exceptions import ValidationError
        cf = CategoryFilter(category=self.parent, characteristic=self.paz,
                            display=CategoryFilter.Display.SELECT_CHECKBOX)  # число + список = нельзя
        with self.assertRaises(ValidationError):
            cf.clean()

    def test_filters_endpoint_returns_effective_config(self):
        CategoryFilter.objects.create(category=self.parent, characteristic=self.paz,
                                      display=CategoryFilter.Display.NUMBER_RANGE,
                                      config={"label": "Размер паза"})
        CategoryFilter.objects.create(category=self.child, characteristic=self.system,
                                      display=CategoryFilter.Display.SELECT_CHECKBOX)
        client = APIClient()
        resp = client.get(f"/api/categories/{self.child.slug}/filters/")
        self.assertEqual(resp.status_code, 200)
        by_code = {f["code"]: f for f in resp.json()}
        self.assertEqual(by_code["paz"]["name"], "Размер паза")     # переопределённая подпись
        self.assertEqual(by_code["paz"]["unit"], "дюйм")
        self.assertEqual(sorted(by_code["system"]["options"]), ["MBT", "Roth"])


class CategoryCharacteristicsWidgetTests(TestCase):
    """Управление составом характеристик категории прямо со страницы категории."""

    def setUp(self):
        self.c1 = Characteristic.objects.create(name="Паз", code="paz", type=Characteristic.Type.NUMBER)
        self.c2 = Characteristic.objects.create(name="Система", code="system", type=Characteristic.Type.SINGLE_SELECT)

    def test_initial_loads_attached_characteristics(self):
        from catalog.admin import CategoryAdminForm
        cat = Category.objects.create(name="Брекеты")
        cat.characteristics.add(self.c1)
        form = CategoryAdminForm(instance=cat)
        initial = list(form.fields["characteristics"].initial)
        self.assertIn(self.c1, initial)
        self.assertNotIn(self.c2, initial)

    def test_save_attaches_selected_characteristics(self):
        from catalog.admin import CategoryAdminForm
        form = CategoryAdminForm(data={
            "name": "Брекеты", "slug": "", "parent": "",
            "characteristics": [self.c1.pk, self.c2.pk],
        })
        self.assertTrue(form.is_valid(), form.errors)
        cat = form.save()
        self.assertEqual(
            set(cat.characteristics.values_list("pk", flat=True)),
            {self.c1.pk, self.c2.pk},
        )

    def test_reverse_relation_consistent(self):
        # привязка со стороны категории видна и со стороны характеристики
        cat = Category.objects.create(name="Дуги")
        cat.characteristics.set([self.c1])
        self.assertIn(cat, self.c1.categories.all())


class CategoryRebuildButtonTests(TestCase):
    """Кнопка «Пересобрать витрину» на списке категорий доступна и не падает."""

    def test_category_rebuild_endpoint_redirects(self):
        User.objects.create_superuser("cat_rb", "c@example.com", "pw12345")
        self.client.force_login(User.objects.get(username="cat_rb"))
        with mock.patch.dict(os.environ, {"GITHUB_DISPATCH_TOKEN": "", "STOREFRONT_REPO": ""}):
            resp = self.client.post("/admin/catalog/category/rebuild/")
        self.assertEqual(resp.status_code, 302)  # редирект обратно на список категорий


class CharacteristicSortableAdminTests(TestCase):
    """Список характеристик стал перетаскиваемым (SortableAdminMixin) — страница грузится."""

    def test_characteristic_changelist_loads(self):
        Characteristic.objects.create(name="Цвет", code="color", type=Characteristic.Type.SINGLE_SELECT)
        User.objects.create_superuser("char_adm", "ch@example.com", "pw12345")
        self.client.force_login(User.objects.get(username="char_adm"))
        resp = self.client.get("/admin/catalog/characteristic/")
        self.assertEqual(resp.status_code, 200)


class ServerMetricsTests(TestCase):
    """Панель метрик: эндпоинт только для персонала и отдаёт корректный JSON."""

    def test_metrics_requires_staff(self):
        resp = self.client.get(reverse("server_metrics"))
        self.assertEqual(resp.status_code, 302)  # аноним -> редирект на логин админки

    def test_metrics_returns_json_for_staff(self):
        User.objects.create_superuser("metrics_adm", "m@example.com", "pw12345")
        self.client.force_login(User.objects.get(username="metrics_adm"))
        resp = self.client.get(reverse("server_metrics"))
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp["Content-Type"], "application/json")
        self.assertIn("available", resp.json())


class BrandLogoSvgTests(TestCase):
    """Логотип бренда принимает SVG (поле FileField), но не «грязный» SVG и не чужие расширения."""

    CLEAN_SVG = b'<svg xmlns="http://www.w3.org/2000/svg" width="10" height="10"><rect width="10" height="10"/></svg>'
    SCRIPT_SVG = b'<svg xmlns="http://www.w3.org/2000/svg"><script>alert(1)</script></svg>'

    def _form(self, filename, content):
        from catalog.admin import BrandAdminForm
        upload = SimpleUploadedFile(filename, content, content_type="image/svg+xml")
        return BrandAdminForm(data={"name": "Acme", "slug": "", "description": ""},
                              files={"logo": upload})

    def test_accepts_clean_svg(self):
        self.assertTrue(self._form("logo.svg", self.CLEAN_SVG).is_valid())

    def test_rejects_svg_with_script(self):
        form = self._form("logo.svg", self.SCRIPT_SVG)
        self.assertFalse(form.is_valid())
        self.assertIn("logo", form.errors)

    def test_rejects_disallowed_extension(self):
        form = self._form("logo.txt", b"hello")
        self.assertFalse(form.is_valid())
        self.assertIn("logo", form.errors)


@override_settings(MEDIA_ROOT=tempfile.mkdtemp())
class TinymceUploadTests(TestCase):
    """Загрузка изображений из TinyMCE: staff грузит картинку, остальное отклоняется."""

    @classmethod
    def setUpTestData(cls):
        cls.staff = User.objects.create_user("ed", "e@e.com", "pw", is_staff=True)

    def _png(self, name="кар тинка.png"):
        from PIL import Image
        buf = BytesIO()
        Image.new("RGB", (2, 2), "red").save(buf, "PNG")
        return SimpleUploadedFile(name, buf.getvalue(), content_type="image/png")

    def test_anonymous_blocked(self):
        r = self.client.post(reverse("tinymce_upload"), {"file": self._png()})
        self.assertIn(r.status_code, (302, 403))

    def test_staff_uploads_png(self):
        self.client.force_login(self.staff)
        r = self.client.post(reverse("tinymce_upload"), {"file": self._png()})
        self.assertEqual(r.status_code, 200)
        loc = r.json().get("location", "")
        self.assertTrue(loc.startswith("http"))
        self.assertIn("/media/uploads/tinymce/", loc)

    def test_rejects_non_image(self):
        self.client.force_login(self.staff)
        bad = SimpleUploadedFile("a.txt", b"hello", content_type="text/plain")
        r = self.client.post(reverse("tinymce_upload"), {"file": bad})
        self.assertEqual(r.status_code, 400)

    def test_rejects_script_svg(self):
        self.client.force_login(self.staff)
        svg = SimpleUploadedFile(
            "x.svg",
            b"<svg xmlns='http://www.w3.org/2000/svg'><script>alert(1)</script></svg>",
            content_type="image/svg+xml",
        )
        r = self.client.post(reverse("tinymce_upload"), {"file": svg})
        self.assertEqual(r.status_code, 400)


class ClientModelTests(TestCase):
    def test_password_set_and_check(self):
        from catalog.clients import Client
        c = Client(email="a@example.com")
        c.set_password("secret123")
        c.save()
        self.assertTrue(c.check_password("secret123"))
        self.assertFalse(c.check_password("wrong"))

    def test_channels_by_approved_memberships(self):
        from catalog.clients import Client, ClientMembership, LegalEntity
        c = Client.objects.create(email="x@example.com")
        self.assertEqual(c.channels(), {"individuals"})
        clinic = LegalEntity.objects.create(inn="7700000000", name="Клиника", segment="clinics")
        ClientMembership.objects.create(client=c, legal_entity=clinic, status="approved")
        self.assertEqual(c.channels(), {"individuals", "clinics"})
        # неподтверждённая привязка канал не даёт
        dist = LegalEntity.objects.create(inn="7700000001", name="Дистр", segment="distributors")
        ClientMembership.objects.create(client=c, legal_entity=dist, status="pending")
        self.assertEqual(c.channels(), {"individuals", "clinics"})


class ContentContractTests(TestCase):
    """Инварианты контентного контракта для внешней интеграции (напр. Ensi):
    единый ключ sku, выборка категории с потомками, целостность серии,
    стабильная форма публичной карточки."""

    def setUp(self):
        self.client = APIClient()
        self.root = Category.objects.create(name="Хирургия и имплантология")
        self.child = Category.objects.create(name="Мембраны", parent=self.root)
        self.p = Product.objects.create(name="CollOss Мембрана 15×20", category=self.child)

    def test_lookup_by_sku_action(self):
        r = self.client.get(f"/api/products/by-sku/{self.p.sku}/")
        self.assertEqual(r.status_code, 200)
        self.assertEqual(r.data["sku"], self.p.sku)
        self.assertEqual(r.data["slug"], self.p.slug)

    def test_list_filter_sku(self):
        r = self.client.get(f"/api/products/?sku={self.p.sku}")
        self.assertEqual(r.status_code, 200)
        slugs = [item["slug"] for item in r.data["results"]]
        self.assertIn(self.p.slug, slugs)

    def test_category_filter_includes_descendants(self):
        # товар лежит в подкатегории — фильтр по корню обязан его вернуть (MPTT-потомки)
        r = self.client.get(f"/api/products/?category={self.root.id}")
        self.assertEqual(r.status_code, 200)
        slugs = [item["slug"] for item in r.data["results"]]
        self.assertIn(self.p.slug, slugs)

    def test_detail_shape_has_content_keys(self):
        r = self.client.get(f"/api/products/{self.p.slug}/")
        self.assertEqual(r.status_code, 200)
        for key in ("sku", "name", "slug", "category", "characteristics",
                    "documents", "images", "group", "audiences", "directions"):
            self.assertIn(key, r.data)


class GroupLevelIntegrityTests(TestCase):
    """Уровень серии должен принадлежать выбранной серии товара."""

    def setUp(self):
        from catalog.models import ProductGroup, GroupLevel
        self.cat = Category.objects.create(name="Импланты")
        self.g1 = ProductGroup.objects.create(name="Серия A")
        self.g2 = ProductGroup.objects.create(name="Серия B")
        self.l1 = GroupLevel.objects.create(group=self.g1, name="Диаметр")
        self.l2 = GroupLevel.objects.create(group=self.g2, name="Длина")

    def test_level_from_other_series_rejected(self):
        from django.core.exceptions import ValidationError
        p = Product(name="Имплант X", category=self.cat, group=self.g1, group_level=self.l2)
        with self.assertRaises(ValidationError):
            p.clean()

    def test_level_without_series_rejected(self):
        from django.core.exceptions import ValidationError
        p = Product(name="Имплант Y", category=self.cat, group=None, group_level=self.l1)
        with self.assertRaises(ValidationError):
            p.clean()

    def test_matching_level_ok(self):
        p = Product(name="Имплант Z", category=self.cat, group=self.g1, group_level=self.l1)
        p.clean()  # не должно бросать


def _make_zip(files):
    """Собрать zip в памяти: files = {arcname: content_str}."""
    import zipfile
    bio = BytesIO()
    with zipfile.ZipFile(bio, "w") as z:
        for name, content in files.items():
            z.writestr(name, content)
    bio.seek(0)
    return bio.getvalue()


_TINCAN = (
    '<?xml version="1.0"?>'
    '<tincan xmlns="http://projecttincan.com/tincan.xsd">'
    '<activities><activity><launch>res/index.html</launch></activity></activities>'
    '</tincan>'
)


class CoursePackageTests(TestCase):
    """Пакет слайдов iSpring/xAPI: безопасная распаковка, поиск точки входа,
    отдача embed_url в API."""

    def setUp(self):
        import shutil
        from catalog.education import Course
        self.tmp = tempfile.mkdtemp()
        ov = override_settings(MEDIA_ROOT=self.tmp)
        ov.enable()
        self.addCleanup(ov.disable)
        self.addCleanup(lambda: shutil.rmtree(self.tmp, ignore_errors=True))
        self.api = APIClient()
        self.cat = Category.objects.create(name="Обучающие товары")
        self.course = Course.objects.create(title="Курс А", slug="kurs-a")

    def _module_with_zip(self, files, name="course.zip"):
        from catalog.education import CourseModule
        m = CourseModule.objects.create(course=self.course, kind="slides", title="Слайды")
        m.package.save(name, SimpleUploadedFile(name, _make_zip(files), "application/zip"))
        m.refresh_from_db()
        return m

    def test_unpack_finds_entry_via_tincan(self):
        m = self._module_with_zip({
            "course/tincan.xml": _TINCAN,
            "course/res/index.html": "<!doctype html><html><body>ok</body></html>",
            "course/res/lms.js": "// player",
        })
        self.assertTrue(m.entry_path.endswith("course/res/index.html"), m.entry_path)
        # файл действительно распакован на диск
        import os as _os
        self.assertTrue(_os.path.isfile(_os.path.join(self.tmp, m.entry_path)))

    def test_zip_slip_rejected(self):
        import os as _os
        from catalog.education_packages import module_dir
        m = self._module_with_zip({
            "../evil.txt": "pwned",
            "res/index.html": "ok",
            "res/lms.js": "//",
        }, name="evil.zip")
        # распаковка отклонена: точки входа нет, файл наружу не записан
        self.assertEqual(m.entry_path, "")
        outside = _os.path.join(_os.path.dirname(module_dir(m.pk)), "evil.txt")
        self.assertFalse(_os.path.exists(outside))

    def test_api_returns_embed_url(self):
        from catalog.education import CourseModule
        product = Product.objects.create(name="Товар с обучением", category=self.cat)
        self.course.products.add(product)
        m = self._module_with_zip({
            "course/tincan.xml": _TINCAN,
            "course/res/index.html": "<!doctype html>ok",
            "course/res/lms.js": "//",
        })
        r = self.api.get(f"/api/products/{product.slug}/")
        self.assertEqual(r.status_code, 200)
        mods = r.data["courses"][0]["modules"]
        slides = [x for x in mods if x["kind"] == "slides"][0]
        self.assertIn("embed_url", slides)
        self.assertTrue(slides["embed_url"].endswith("res/index.html"), slides["embed_url"])
        self.assertTrue(slides["embed_url"].startswith("http"))


class ImageDerivativeTests(TestCase):
    """Уменьшенные копии: генерация, вписывание, отсутствие апскейла, URL в API."""

    def setUp(self):
        import shutil
        self.tmp = tempfile.mkdtemp()
        ov = override_settings(MEDIA_ROOT=self.tmp)
        ov.enable()
        self.addCleanup(ov.disable)
        self.addCleanup(lambda: shutil.rmtree(self.tmp, ignore_errors=True))
        self.api = APIClient()
        self.cat = Category.objects.create(name="Товары с фото")
        self.product = Product.objects.create(name="Фото-товар", category=self.cat)

    def _png(self, w=2000, h=1500):
        from PIL import Image
        buf = BytesIO()
        Image.new("RGB", (w, h), (180, 90, 40)).save(buf, format="PNG")
        return buf.getvalue()

    def _add_image(self, w=2000, h=1500, name="orig.png"):
        from catalog.models import ProductImage
        img = ProductImage(product=self.product)
        img.image.save(name, SimpleUploadedFile(name, self._png(w, h), "image/png"))
        img.refresh_from_db()
        return img

    def _dims(self, filefield):
        from PIL import Image
        import os as _os
        with Image.open(_os.path.join(self.tmp, filefield.name)) as im:
            return im.size

    def test_three_copies_created_and_fit(self):
        img = self._add_image(2000, 1500)
        self.assertTrue(img.thumb and img.card and img.main)
        self.assertLessEqual(max(self._dims(img.thumb)), 160)
        self.assertLessEqual(max(self._dims(img.card)), 400)
        self.assertLessEqual(max(self._dims(img.main)), 1200)
        # пропорции сохранены (не квадрат)
        w, h = self._dims(img.main)
        self.assertAlmostEqual(w / h, 2000 / 1500, places=1)

    def test_no_upscale(self):
        img = self._add_image(120, 90, name="small.png")
        self.assertLessEqual(max(self._dims(img.main)), 120)

    def test_gallery_and_list_urls(self):
        self._add_image()
        r = self.api.get(f"/api/products/{self.product.slug}/")
        gallery = r.data["images"][0]
        for key in ("thumb", "main", "original", "alt", "order"):
            self.assertIn(key, gallery)
        self.assertTrue(gallery["thumb"].startswith("http"))
        self.assertTrue(gallery["main"].startswith("http"))
        # список каталога: thumbnail = копия card
        rl = self.api.get(f"/api/products/?sku={self.product.sku}")
        item = [x for x in rl.data["results"] if x["slug"] == self.product.slug][0]
        self.assertTrue(item["thumbnail"].startswith("http"))


class ProductActiveTests(TestCase):
    """is_active: неактивные товары не отдаются публичным API."""

    def setUp(self):
        self.api = APIClient()
        self.cat = Category.objects.create(name="Категория A")
        self.active = Product.objects.create(name="Активный", category=self.cat, is_active=True)
        self.hidden = Product.objects.create(name="Скрытый", category=self.cat, is_active=False)

    def test_list_excludes_inactive(self):
        r = self.api.get("/api/products/")
        slugs = [x["slug"] for x in r.data["results"]]
        self.assertIn(self.active.slug, slugs)
        self.assertNotIn(self.hidden.slug, slugs)

    def test_detail_of_inactive_is_404(self):
        r = self.api.get(f"/api/products/{self.hidden.slug}/")
        self.assertEqual(r.status_code, 404)

    def test_active_detail_ok(self):
        r = self.api.get(f"/api/products/{self.active.slug}/")
        self.assertEqual(r.status_code, 200)


class MediaDedupTests(TestCase):
    """Content-addressed хранение + удаление по счётчику ссылок."""

    def setUp(self):
        import shutil
        self.tmp = tempfile.mkdtemp()
        ov = override_settings(MEDIA_ROOT=self.tmp)
        ov.enable()
        self.addCleanup(ov.disable)
        self.addCleanup(lambda: shutil.rmtree(self.tmp, ignore_errors=True))
        self.cat = Category.objects.create(name="Дедуп")
        self.p1 = Product.objects.create(name="Товар 1", category=self.cat)
        self.p2 = Product.objects.create(name="Товар 2", category=self.cat)

    def _png(self, color=(1, 2, 3)):
        from PIL import Image
        buf = BytesIO()
        Image.new("RGB", (300, 200), color).save(buf, format="PNG")
        return buf.getvalue()

    def _add(self, product, data, name="orig.png"):
        from catalog.models import ProductImage
        img = ProductImage(product=product)
        img.image.save(name, SimpleUploadedFile(name, data, "image/png"))
        img.refresh_from_db()
        return img

    @property
    def _storage(self):
        from catalog.storage import product_image_storage
        return product_image_storage

    def test_identical_files_are_shared(self):
        data = self._png()
        i1 = self._add(self.p1, data)
        i2 = self._add(self.p2, data)
        # одинаковые байты → один и тот же путь по всем четырём файлам
        self.assertEqual(i1.image.name, i2.image.name)
        self.assertEqual(i1.thumb.name, i2.thumb.name)
        self.assertEqual(i1.main.name, i2.main.name)

    def test_delete_one_keeps_shared_file(self):
        data = self._png()
        i1 = self._add(self.p1, data)
        self._add(self.p2, data)
        shared = i1.image.name
        i1.delete()
        # второй товар всё ещё ссылается → файл ОБЯЗАН остаться
        self.assertTrue(self._storage.exists(shared))

    def test_delete_last_owner_removes_file(self):
        data = self._png()
        i1 = self._add(self.p1, data)
        i2 = self._add(self.p2, data)
        shared = i1.image.name
        i1.delete()
        i2.delete()
        self.assertFalse(self._storage.exists(shared))

    def test_replace_removes_orphaned_original(self):
        from catalog.models import ProductImage
        i = self._add(self.p1, self._png(color=(1, 2, 3)))
        old_name = i.image.name
        inst = ProductImage.objects.get(pk=i.pk)  # from_db → запомнит старые имена
        inst.image.save("new.png", SimpleUploadedFile("new.png", self._png(color=(9, 9, 9)), "image/png"))
        inst.refresh_from_db()
        self.assertNotEqual(inst.image.name, old_name)
        self.assertFalse(self._storage.exists(old_name))       # старый осиротел → удалён
        self.assertTrue(self._storage.exists(inst.image.name))  # новый на месте

    def test_replace_keeps_file_still_shared(self):
        from catalog.models import ProductImage
        data = self._png()
        i1 = self._add(self.p1, data)
        self._add(self.p2, data)        # p2 делит те же файлы
        shared = i1.image.name
        inst = ProductImage.objects.get(pk=i1.pk)
        inst.image.save("new.png", SimpleUploadedFile("new.png", self._png(color=(7, 7, 7)), "image/png"))
        # p1 сменил оригинал, но p2 всё ещё ссылается на старый файл → он остаётся
        self.assertTrue(self._storage.exists(shared))

    def test_migrate_to_cas_idempotent(self):
        from catalog.image_derivatives import migrate_to_cas
        from catalog.models import ProductImage
        i = self._add(self.p1, self._png())
        inst = ProductImage.objects.get(pk=i.pk)
        self.assertEqual(migrate_to_cas(inst), 0)  # файлы уже под хеш-именами


class AudienceFromDirectionsTests(TestCase):
    """Аудитории товара выводятся из аудиторий его направлений."""

    def setUp(self):
        from catalog.models import Audience, Direction
        # изолированные slug'и, чтобы не конфликтовать с засеянными миграцией 0008
        self.a_stom = Audience.objects.create(name="Т-Стоматолог", slug="t-stom")
        self.a_tech = Audience.objects.create(name="Т-Техник", slug="t-tech")
        self.a_med = Audience.objects.create(name="Т-Медицина", slug="t-med")
        self.d_terapiya = Direction.objects.create(name="Т-Терапия", slug="t-terapiya", audience=self.a_stom)
        self.d_cadcam = Direction.objects.create(name="Т-CADCAM", slug="t-cadcam", audience=self.a_tech)
        self.cat = Category.objects.create(name="Кат")

    def _aud_slugs(self, product):
        return set(product.audiences.values_list("slug", flat=True))

    def test_audience_added_from_direction(self):
        p = Product.objects.create(name="Товар 1", category=self.cat)
        p.directions.add(self.d_terapiya)
        self.assertEqual(self._aud_slugs(p), {"t-stom"})

    def test_multiple_directions_union(self):
        p = Product.objects.create(name="Товар 2", category=self.cat)
        p.directions.set([self.d_terapiya, self.d_cadcam])
        self.assertEqual(self._aud_slugs(p), {"t-stom", "t-tech"})

    def test_removing_direction_updates_audiences(self):
        p = Product.objects.create(name="Товар 3", category=self.cat)
        p.directions.set([self.d_terapiya, self.d_cadcam])
        p.directions.remove(self.d_cadcam)
        self.assertEqual(self._aud_slugs(p), {"t-stom"})

    def test_manual_audience_kept_without_directions(self):
        p = Product.objects.create(name="Перчатки", category=self.cat)
        p.audiences.add(self.a_med)  # ручной выбор, направлений нет
        self.assertEqual(self._aud_slugs(p), {"t-med"})


class CompatibilitySystemTests(TestCase):
    """Фасет «Система совместимости»: API-выдача, признак, фильтр, валидация."""

    def setUp(self):
        from catalog.models import CompatibilitySystem
        self.api = APIClient()
        self.cat = Category.objects.create(name="Абатменты")
        self.sys_straumann = CompatibilitySystem.objects.create(
            name="Straumann BLT", slug="implantatsiya-straumann-blt", group="Имплантация", order=1)
        self.sys_dentium = CompatibilitySystem.objects.create(
            name="Dentium SuperLine", slug="implantatsiya-dentium-superline", group="Имплантация", order=2)
        # системозависимый товар
        self.p_fit = Product.objects.create(name="Абатмент совместимый", category=self.cat,
                                             fitment_type="compatible")
        self.p_fit.compatibility_systems.add(self.sys_straumann)
        # системонезависимый товар
        self.p_plain = Product.objects.create(name="Пломбировочный материал", category=self.cat)

    def test_detail_outputs_systems_and_fitment(self):
        r = self.api.get(f"/api/products/{self.p_fit.slug}/")
        self.assertEqual(r.status_code, 200)
        slugs = [s["slug"] for s in r.data["systems"]]
        self.assertEqual(slugs, ["implantatsiya-straumann-blt"])
        self.assertEqual(r.data["fitment"], "compatible")

    def test_list_outputs_system_slugs(self):
        r = self.api.get(f"/api/products/?sku={self.p_fit.sku}")
        item = [x for x in r.data["results"] if x["slug"] == self.p_fit.slug][0]
        self.assertIn("implantatsiya-straumann-blt", item["systems"])
        self.assertEqual(item["fitment"], "compatible")

    def test_fitment_hidden_without_systems(self):
        r = self.api.get(f"/api/products/{self.p_plain.slug}/")
        self.assertEqual(r.data["systems"], [])
        self.assertIsNone(r.data["fitment"])

    def test_filter_by_system(self):
        r = self.api.get("/api/products/?system=implantatsiya-straumann-blt")
        slugs = [x["slug"] for x in r.data["results"]]
        self.assertIn(self.p_fit.slug, slugs)
        self.assertNotIn(self.p_plain.slug, slugs)

    def test_systems_reference_endpoint(self):
        r = self.api.get("/api/systems/")
        results = r.data["results"] if isinstance(r.data, dict) and "results" in r.data else r.data
        slugs = [s["slug"] for s in results]
        self.assertIn("implantatsiya-straumann-blt", slugs)
        self.assertIn("implantatsiya-dentium-superline", slugs)

    def test_write_requires_fitment_when_system_set(self):
        from catalog.serializers import ProductWriteSerializer
        bad = ProductWriteSerializer(data={
            "name": "Новый абатмент", "category": self.cat.id,
            "compatibility_systems": [self.sys_straumann.id]})
        self.assertFalse(bad.is_valid())
        self.assertIn("fitment_type", bad.errors)
        good = ProductWriteSerializer(data={
            "name": "Новый абатмент 2", "category": self.cat.id,
            "compatibility_systems": [self.sys_straumann.id], "fitment_type": "original"})
        self.assertTrue(good.is_valid(), good.errors)


class VariantColorTests(TestCase):
    """Цвет варианта: нормализация, валидация, выдача в API, Excel round-trip."""

    def setUp(self):
        from catalog.models import ProductGroup
        self.api = APIClient()
        self.cat = Category.objects.create(name="Керамика")
        self.grp = ProductGroup.objects.create(name="SD Ceram Stain", slug="sd-ceram-stain")
        self.p_beige = Product.objects.create(
            name="SD Ceram Beige Fluor", category=self.cat, group=self.grp,
            variant_label="Beige", variant_color="#C8A165")
        self.p_kit = Product.objects.create(
            name="SD Ceram Kit", category=self.cat, group=self.grp, variant_label="Kit")

    def test_normalization_forms(self):
        from catalog.utils import normalize_hex_color
        self.assertEqual(normalize_hex_color("c8a165"), "#C8A165")
        self.assertEqual(normalize_hex_color("  #c8a165 "), "#C8A165")
        self.assertEqual(normalize_hex_color("#CA6"), "#CCAA66")
        self.assertIsNone(normalize_hex_color("не цвет"))
        self.assertIsNone(normalize_hex_color("rgb(1,2,3)"))
        self.assertIsNone(normalize_hex_color(""))

    def test_saved_in_canonical_form(self):
        p = Product.objects.create(name="Тест цвет", category=self.cat, variant_color="a1b2c3")
        p.refresh_from_db()
        self.assertEqual(p.variant_color, "#A1B2C3")

    def test_validator_rejects_garbage(self):
        from django.core.exceptions import ValidationError
        from catalog.utils import validate_hex_color
        with self.assertRaises(ValidationError):
            validate_hex_color("зелёненький")
        validate_hex_color("")          # пусто — допустимо
        validate_hex_color("#C8A165")   # корректно

    def test_api_variant_exposes_color(self):
        r = self.api.get(f"/api/products/{self.p_beige.slug}/")
        self.assertEqual(r.status_code, 200)
        variants = {v["label"]: v for v in r.data["group"]["variants"]}
        self.assertEqual(variants["Beige"]["color"], "#C8A165")
        # у товара без цвета (набор) поле есть, но пустое
        self.assertIsNone(variants["Kit"]["color"])

    def test_excel_roundtrip_color(self):
        from catalog.catalog_io import build_export_workbook, import_workbook
        from io import BytesIO
        wb = build_export_workbook(Product.objects.filter(pk=self.p_beige.pk))
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        report = import_workbook(buf)
        self.p_beige.refresh_from_db()
        self.assertEqual(self.p_beige.variant_color, "#C8A165")
        self.assertEqual(report.get("errors", []), [])

    def test_excel_import_warns_on_bad_color(self):
        from catalog.catalog_io import build_export_workbook, import_workbook
        from io import BytesIO
        wb = build_export_workbook(Product.objects.filter(pk=self.p_beige.pk))
        ws = wb.active
        headers = [c.value for c in ws[1]]           # строка 1 — машинные ключи
        col = headers.index("variant_color") + 1
        ws.cell(row=3, column=col, value="бежевенький")   # данные идут с 3-й строки
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        report = import_workbook(buf)
        self.p_beige.refresh_from_db()
        self.assertEqual(self.p_beige.variant_color, "#C8A165")  # старое значение уцелело
        self.assertTrue(any("не похоже на цвет" in w for w in report["warnings"]), report["warnings"])


class GroupEditorColorTests(TestCase):
    """Редактор группировок: цвет варианта читается в состоянии и сохраняется."""

    def setUp(self):
        from django.contrib.auth.models import User
        from django.test import Client as DjangoClient  # в проекте есть своя модель Client
        from catalog.models import ProductGroup
        self.staff = User.objects.create_superuser("ge_admin", "a@e.com", "pw12345!")
        self.client = DjangoClient()
        self.client.force_login(self.staff)
        self.cat = Category.objects.create(name="Керамика GE")
        self.grp = ProductGroup.objects.create(name="Серия GE", slug="seriya-ge")
        self.p = Product.objects.create(name="Stain Beige", category=self.cat,
                                        group=self.grp, variant_label="Beige",
                                        variant_color="#C8A165")

    def _state(self):
        return self.client.get(f"/admin/catalog/productgroup/{self.grp.pk}/editor/state/")

    def _save(self, payload):
        import json
        return self.client.post(f"/admin/catalog/productgroup/{self.grp.pk}/editor/save/",
                                data=json.dumps(payload), content_type="application/json")

    def test_state_includes_color(self):
        r = self._state()
        self.assertEqual(r.status_code, 200)
        member = r.json()["members"][0]
        self.assertEqual(member["variant_color"], "#C8A165")

    def test_save_updates_color_and_normalizes(self):
        r = self._save({"members": [{"id": self.p.pk, "variant_label": "Beige",
                                     "variant_color": "a1b2c3",
                                     "group_order": 0, "group_level": None}]})
        self.assertEqual(r.status_code, 200)
        self.p.refresh_from_db()
        self.assertEqual(self.p.variant_color, "#A1B2C3")

    def test_save_empty_clears_color(self):
        self._save({"members": [{"id": self.p.pk, "variant_label": "Beige",
                                 "variant_color": "", "group_order": 0, "group_level": None}]})
        self.p.refresh_from_db()
        self.assertEqual(self.p.variant_color, "")

    def test_save_garbage_keeps_previous(self):
        self._save({"members": [{"id": self.p.pk, "variant_label": "Beige",
                                 "variant_color": "бежевенький",
                                 "group_order": 0, "group_level": None}]})
        self.p.refresh_from_db()
        self.assertEqual(self.p.variant_color, "#C8A165")
