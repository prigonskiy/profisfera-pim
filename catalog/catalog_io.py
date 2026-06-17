"""
Общая логика выгрузки/загрузки каталога в xlsx.

Сейчас здесь живёт ЭКСПОРТ; столбцы и форматирование значений вынесены сюда же,
чтобы будущий импорт читал ровно тот же формат (DRY).

Формат файла:
  • лист «Товары» — по строке на товар;
  • первый столбец «Артикул PIM» (sku) — КЛЮЧ сопоставления при импорте;
  • многозначные поля (аудитории, направления, мультисписки) — через разделитель " | ";
  • заголовки столбцов характеристик — читаемые имена характеристик;
  • лист «Справка» — словарь кодов/имён/типов и допустимых значений.
"""
from io import BytesIO
from decimal import Decimal, InvalidOperation

from django.db import transaction
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .models import (
    Audience,
    Brand,
    Category,
    Characteristic,
    CharacteristicOption,
    Direction,
    Product,
    ProductAttributeValue,
)

MULTI_SEP = " | "

# Базовые (некатегорийные) столбцы: (заголовок, ключ, ширина)
CORE_COLUMNS = [
    ("Артикул PIM", "sku", 16),
    ("Код сопоставления (ERP)", "external_id", 20),
    ("Название", "name", 40),
    ("Slug", "slug", 22),
    ("Бренд", "brand", 22),
    ("Категория", "category_path", 32),
    ("Категория (slug)", "category_slug", 24),
    ("Артикул", "manufacturer_sku", 18),
    ("GTIN", "gtin", 16),
    ("Код ТН ВЭД", "tnved_code", 14),
    ("Страна (ISO)", "country", 12),
    ("Краткое описание", "short_description", 45),
    ("Полное описание (HTML)", "full_description", 60),
    ("Ширина брутто, мм", "gross_width_mm", 16),
    ("Высота брутто, мм", "gross_height_mm", 16),
    ("Глубина брутто, мм", "gross_depth_mm", 16),
    ("Масса брутто, кг", "gross_weight_kg", 16),
    ("Аудитории", "audiences", 28),
    ("Направления", "directions", 32),
]

HEADER_FILL = PatternFill("solid", fgColor="0E2A31")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
KEY_FILL = PatternFill("solid", fgColor="0A6E73")
ID_FILL = PatternFill("solid", fgColor="EEF3F4")
ID_FONT = Font(name="Arial", size=8, color="7A949B", italic=True)
BASE_FONT = Font(name="Arial", size=10)
THIN = Side(style="thin", color="D9E2E4")
BORDER = Border(bottom=THIN, right=THIN)


def _num(d):
    return float(d) if d is not None else None


def _category_map():
    """id -> (name, parent_id) для быстрого построения пути и потомков без доп. запросов."""
    return {c.id: (c.name, c.parent_id) for c in Category.objects.all()}


def category_path(cat_id, cmap):
    parts, seen = [], set()
    while cat_id and cat_id not in seen:
        seen.add(cat_id)
        name, parent_id = cmap.get(cat_id, (None, None))
        if name is None:
            break
        parts.append(name)
        cat_id = parent_id
    return " / ".join(reversed(parts))


def descendant_ids(root_id, cmap):
    """root + все потомки (по карте категорий)."""
    children = {}
    for cid, (_, pid) in cmap.items():
        children.setdefault(pid, []).append(cid)
    out, stack = set(), [root_id]
    while stack:
        cur = stack.pop()
        if cur in out:
            continue
        out.add(cur)
        stack.extend(children.get(cur, []))
    return out


def category_with_descendants(root_id):
    """Множество id: категория + все её потомки (для экспорта по категории)."""
    return descendant_ids(root_id, _category_map())


def category_labels():
    """[(id, 'Родитель / Категория'), ...] отсортировано по пути — для выпадающего списка."""
    cmap = _category_map()
    items = [(cid, category_path(cid, cmap)) for cid in cmap]
    items.sort(key=lambda x: x[1])
    return items


def cell_for_value(typed_value, char_type):
    """Типизированное значение характеристики -> значение ячейки (чистая функция)."""
    if typed_value is None or typed_value == "":
        return None
    if char_type == "multi_select":
        return MULTI_SEP.join(str(v) for v in typed_value)
    if char_type == "boolean":
        return "Да" if typed_value else "Нет"
    if char_type == "number":
        return _num(typed_value)
    return str(typed_value)


def _collect_characteristics(products, include_category_chars):
    """Глобальные характеристики (всегда) + категорийные (для категорий выбранных товаров)."""
    chars = list(Characteristic.objects.filter(is_global=True).order_by("name"))
    if include_category_chars:
        cat_ids = {p.category_id for p in products if p.category_id}
        if cat_ids:
            cat_chars = (
                Characteristic.objects.filter(is_global=False, categories__in=cat_ids)
                .distinct()
                .order_by("name")
            )
            chars += list(cat_chars)
    return chars


def _core_value(product, key, cmap):
    if key == "brand":
        return product.brand.name if product.brand_id else None
    if key == "category_path":
        return category_path(product.category_id, cmap) if product.category_id else None
    if key == "category_slug":
        return product.category.slug if product.category_id else None
    if key == "country":
        c = product.country_of_origin
        return c.code if c else None
    if key in ("gross_width_mm", "gross_height_mm", "gross_depth_mm", "gross_weight_kg"):
        return _num(getattr(product, key))
    if key == "audiences":
        return MULTI_SEP.join(a.name for a in product.audiences.all()) or None
    if key == "directions":
        return MULTI_SEP.join(d.name for d in product.directions.all()) or None
    return getattr(product, key) or None


def build_export_workbook(queryset, include_category_chars=False):
    products = list(
        queryset.select_related("brand", "category").prefetch_related(
            "audiences", "directions", "attribute_values__characteristic", "attribute_values__value_options"
        )
    )
    cmap = _category_map()
    chars = _collect_characteristics(products, include_category_chars)

    wb = Workbook()
    ws = wb.active
    ws.title = "Товары"

    core_labels = [h for h, _, _ in CORE_COLUMNS]
    core_ids = [key for _, key, _ in CORE_COLUMNS]
    id_row = core_ids + ["char:" + c.code for c in chars]      # строка 1 — машинные ключи
    label_row = core_labels + [c.name for c in chars]          # строка 2 — читаемые заголовки
    ncols = len(id_row)

    ws.append(id_row)
    ws.append(label_row)
    for col in range(1, ncols + 1):
        c1 = ws.cell(row=1, column=col)
        c1.font = ID_FONT
        c1.fill = ID_FILL
        c1.alignment = Alignment(vertical="center")
        c2 = ws.cell(row=2, column=col)
        c2.font = HEADER_FONT
        c2.fill = KEY_FILL if col == 1 else HEADER_FILL
        c2.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 14
    ws.row_dimensions[2].height = 30

    first_data_row = 3
    for i, product in enumerate(products):
        # значения характеристик товара: char_id -> av
        av_by_char = {av.characteristic_id: av for av in product.attribute_values.all()}
        row = [_core_value(product, key, cmap) for _, key, _ in CORE_COLUMNS]
        for ch in chars:
            av = av_by_char.get(ch.id)
            row.append(cell_for_value(av.value, ch.type) if av else None)
        ws.append(row)
        r = first_data_row + i
        ws.row_dimensions[r].height = 15          # фиксированная высота строки
        for cell in ws[r]:
            cell.font = BASE_FONT
            cell.alignment = Alignment(vertical="center", wrap_text=False)

    for idx, (_, _, width) in enumerate(CORE_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    for idx in range(len(CORE_COLUMNS) + 1, ncols + 1):
        ws.column_dimensions[get_column_letter(idx)].width = 22
    ws.freeze_panes = "B3"      # закрепляем обе строки заголовка и столбец «Артикул PIM»

    _add_legend_sheet(wb, chars)
    return wb


def _add_legend_sheet(wb, chars):
    ws = wb.create_sheet("Справка")
    bold = Font(name="Arial", bold=True, size=10)
    base = Font(name="Arial", size=10)

    def section(title):
        ws.append([title])
        ws[ws.max_row][0].font = bold

    section("Как пользоваться файлом")
    for line in [
        "Строка 1 — машинные идентификаторы (ключи полей и char:<код> характеристик); по ним идёт импорт. Не трогать.",
        "Строка 2 — читаемые заголовки (для человека). Данные — с третьей строки.",
        "Артикул PIM (sku) — внутренний ключ товара; по нему пойдёт сопоставление при импорте. Не меняйте.",
        "Код сопоставления (external_id) — код товара во внешних системах (ERP/Litics); это НЕ ключ PIM.",
        f"Многозначные поля (аудитории, направления, мультисписки) разделяются «{MULTI_SEP.strip()}».",
        "Пустая ячейка = значение не задано. Булевы характеристики: «Да» / «Нет».",
        "Категория сопоставляется по столбцу «Категория (slug)». Страна — ISO-код (например, JP).",
    ]:
        ws.append([line])
        ws[ws.max_row][0].font = base
    ws.append([])

    section("Характеристики (заголовки столбцов)")
    ws.append(["Имя (заголовок)", "Код", "Тип", "Глобальная", "Допустимые значения (для списков)"])
    for c in ws[ws.max_row]:
        c.font = bold
    type_label = dict(Characteristic.Type.choices)
    for ch in chars:
        opts = ""
        if ch.type in ("single_select", "multi_select"):
            opts = MULTI_SEP.join(o.value for o in ch.options.all())
        ws.append([ch.name, ch.code, type_label.get(ch.type, ch.type), "да" if ch.is_global else "нет", opts])
        for c in ws[ws.max_row]:
            c.font = base
    ws.append([])

    section("Аудитории (имя → как писать в ячейке)")
    for a in Audience.objects.all().order_by("order", "name"):
        ws.append([a.name]); ws[ws.max_row][0].font = base
    ws.append([])

    section("Направления (имя · аудитория)")
    for d in Direction.objects.select_related("audience").all():
        ws.append([d.name, d.audience.name if d.audience_id else "—"])
        for c in ws[ws.max_row]:
            c.font = base

    ws.column_dimensions["A"].width = 42
    for col in "BCDE":
        ws.column_dimensions[col].width = 22


def workbook_response(wb, basename):
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    ts = timezone.now().strftime("%Y%m%d_%H%M")
    resp = HttpResponse(
        bio.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{basename}_{ts}.xlsx"'
    return resp


# ===========================================================================
# ИМПОРТ
# ===========================================================================
def _s(v):
    return "" if v is None else str(v).strip()


def _split_multi(v):
    s = _s(v)
    return [p.strip() for p in s.split("|") if p.strip()] if s else []


def _parse_bool(v):
    s = _s(v).lower()
    if s in ("да", "yes", "true", "1", "истина", "y", "+"):
        return True
    if s in ("нет", "no", "false", "0", "ложь", "n", "-"):
        return False
    return None


def _parse_decimal(v):
    s = _s(v).replace(",", ".").replace(" ", "")
    if not s:
        return None
    try:
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return None


def _resolve_category(slug):
    slug = _s(slug)
    return Category.objects.filter(slug=slug).first() if slug else None


def _resolve_brand(name):
    name = _s(name)
    return Brand.objects.get_or_create(name=name)[0] if name else None


def _set_char(product, char, cell):
    """Записать/очистить значение характеристики у товара по содержимому ячейки."""
    existing = ProductAttributeValue.objects.filter(product=product, characteristic=char).first()

    def ensure_av():
        return existing or ProductAttributeValue.objects.create(product=product, characteristic=char)

    if char.type in ("single_select", "multi_select"):
        names = _split_multi(cell) if char.type == "multi_select" else ([_s(cell)] if _s(cell) else [])
        if not names:
            if existing:
                existing.delete()
            return
        av = ensure_av()
        opts = []
        for v in names:
            opt, _ = CharacteristicOption.objects.get_or_create(
                characteristic=char, value=v, defaults={"order": char.options.count()}
            )
            opts.append(opt)
        av.value_options.set(opts)
    elif char.type == "boolean":
        b = _parse_bool(cell)
        if b is None:
            if existing:
                existing.delete()
            return
        av = ensure_av()
        av.value_boolean = b
        av.save(update_fields=["value_boolean"])
    elif char.type == "number":
        num = _parse_decimal(cell)
        if num is None:
            if existing:
                existing.delete()
            return
        av = ensure_av()
        av.value_number = num
        av.save(update_fields=["value_number"])
    else:  # text
        val = _s(cell)
        if not val:
            if existing:
                existing.delete()
            return
        av = ensure_av()
        av.value_text = val
        av.save(update_fields=["value_text"])


def _validate_row(data, headers_present, char_columns, opt_map):
    """Предпроверка значений по типам. Возвращает (ошибки, предупреждения).
    Ошибки отклоняют строку целиком; предупреждения — информируют, но не блокируют."""
    errors, warnings = [], []

    # числовые габариты (брутто)
    for dim, label in (
        ("gross_width_mm", "Ширина брутто"), ("gross_height_mm", "Высота брутто"),
        ("gross_depth_mm", "Глубина брутто"), ("gross_weight_kg", "Масса брутто"),
    ):
        if dim in headers_present:
            s = _s(data.get(dim))
            if s and _parse_decimal(data.get(dim)) is None:
                errors.append(f"«{label}»: ожидается число, получено «{s}»")

    # характеристики
    for header, char in char_columns:
        if header not in headers_present:
            continue
        raw = data.get(header)
        s = _s(raw)
        if not s:
            continue  # пустая ячейка — законная очистка, не ошибка
        if char.type == "boolean" and _parse_bool(raw) is None:
            errors.append(f"«{char.name}»: ожидается Да/Нет, получено «{s}»")
        elif char.type == "number" and _parse_decimal(raw) is None:
            errors.append(f"«{char.name}»: ожидается число, получено «{s}»")
        elif char.type in ("single_select", "multi_select"):
            allowed = opt_map.get(char.id, set())
            values = _split_multi(raw) if char.type == "multi_select" else [s]
            unknown = [v for v in values if v.strip().lower() not in allowed]
            if unknown and allowed:
                warnings.append(
                    f"«{char.name}»: значения не из справочника будут добавлены как новые: {', '.join(unknown)}"
                )
    return errors, warnings


def _apply_row(data, headers_present, char_columns, maps, report, row_no):
    # предпроверка по типам — до любых записей; ошибка отклоняет строку целиком
    errors, warnings = _validate_row(data, headers_present, char_columns, maps.get("opt", {}))
    for w in warnings:
        report["warnings"].append(f"строка {row_no}: {w}")
    if errors:
        raise ValueError("; ".join(errors))

    sku = _s(data.get("sku"))
    ext = _s(data.get("external_id"))
    name = _s(data.get("name"))

    # сопоставление товара
    product, created = None, False
    if sku:
        product = Product.objects.filter(sku=sku).first()
        if product is None:
            raise ValueError(
                f"товар с «Артикул PIM» {sku} не найден. Артикул присваивает система; "
                f"для нового товара оставьте столбец «Артикул PIM» пустым."
            )
    elif ext:
        product = Product.objects.filter(external_id=ext).first()

    if product is None:
        if not name:
            raise ValueError("для нового товара обязательно «Название»")
        cat = _resolve_category(data.get("category_slug"))
        if cat is None:
            raise ValueError("для нового товара нужна существующая «Категория (slug)»")
        product = Product(name=name, category=cat)
        created = True

    # скалярные поля — только те столбцы, что есть в файле
    if "name" in headers_present and name:
        product.name = name
    if "slug" in headers_present and _s(data.get("slug")):
        product.slug = _s(data.get("slug"))
    if "brand" in headers_present:
        product.brand = _resolve_brand(data.get("brand"))
    if "category_slug" in headers_present:
        slug = _s(data.get("category_slug"))
        if slug:
            cat = _resolve_category(slug)
            if cat is None:
                raise ValueError(f"категория со slug «{slug}» не найдена")
            product.category = cat
    if "external_id" in headers_present:
        product.external_id = ext or None
    for f in ("manufacturer_sku", "gtin", "tnved_code", "short_description", "full_description"):
        if f in headers_present:
            setattr(product, f, _s(data.get(f)))
    if "country" in headers_present:
        product.country_of_origin = _s(data.get("country"))
    for dim in ("gross_width_mm", "gross_height_mm", "gross_depth_mm", "gross_weight_kg"):
        if dim in headers_present:
            setattr(product, dim, _parse_decimal(data.get(dim)))

    product.save()

    # таксономия (замена набора)
    if "audiences" in headers_present:
        auds = []
        for n in _split_multi(data.get("audiences")):
            a = maps["aud"].get(n.lower())
            if a:
                auds.append(a)
            else:
                report["warnings"].append(f"строка {row_no}: аудитория «{n}» не найдена — пропущена")
        product.audiences.set(auds)
    if "directions" in headers_present:
        dirs = []
        for n in _split_multi(data.get("directions")):
            d = maps["dir"].get(n.lower())
            if d:
                dirs.append(d)
            else:
                report["warnings"].append(f"строка {row_no}: направление «{n}» не найдено — пропущено")
        product.directions.set(dirs)

    # характеристики
    for header, char in char_columns:
        if header in headers_present:
            _set_char(product, char, data.get(header))

    report["created" if created else "updated"] += 1


def import_workbook(fileobj, dry_run=True):
    """
    Импорт каталога из xlsx нашего формата.
    Сопоставление: по «Артикул PIM» (sku); если он пуст — по external_id; иначе создаётся новый.
    Возвращает отчёт: created / updated / skipped / errors / warnings / applied.
    """
    report = {"created": 0, "updated": 0, "skipped": 0, "errors": [], "warnings": [], "applied": not dry_run}

    wb = load_workbook(fileobj, read_only=True, data_only=True)
    ws = wb["Товары"] if "Товары" in wb.sheetnames else wb.active
    rows = ws.iter_rows(values_only=True)
    try:
        id_row = next(rows)
        next(rows)  # строка 2 — человекочитаемые заголовки, игнорируем
    except StopIteration:
        report["errors"].append((0, "В файле нет строк заголовка."))
        wb.close()
        return report

    headers = [(_s(h) or None) for h in id_row]
    headers_present = {h for h in headers if h}
    if "sku" not in headers_present and "external_id" not in headers_present:
        report["errors"].append((1, "В первой строке нет столбца-идентификатора (sku или external_id)."))
        wb.close()
        return report

    char_by_code = {c.code: c for c in Characteristic.objects.all()}
    char_columns = []
    for h in headers:
        if h and h.startswith("char:"):
            code = h[5:]
            ch = char_by_code.get(code)
            if ch:
                char_columns.append((h, ch))
            else:
                report["warnings"].append(f"неизвестный код характеристики «{h}» — столбец пропущен")
    opt_map = {
        ch.id: {_s(o.value).lower() for o in ch.options.all()}
        for _h, ch in char_columns
        if ch.type in ("single_select", "multi_select")
    }
    maps = {
        "aud": {a.name.strip().lower(): a for a in Audience.objects.all()},
        "dir": {d.name.strip().lower(): d for d in Direction.objects.all()},
        "opt": opt_map,
    }

    with transaction.atomic():
        for i, raw in enumerate(rows, start=3):
            if raw is None or all(c is None or _s(c) == "" for c in raw):
                continue  # пустая строка
            data = {headers[j]: raw[j] for j in range(min(len(headers), len(raw))) if headers[j]}
            try:
                _apply_row(data, headers_present, char_columns, maps, report, i)
            except Exception as e:
                report["errors"].append((i, str(e)))
                report["skipped"] += 1
        if dry_run:
            transaction.set_rollback(True)

    wb.close()
    return report
